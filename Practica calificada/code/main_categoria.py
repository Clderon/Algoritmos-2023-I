# from openpyxl import Workbook, load_workbook
import pandas as pd
import openpyxl
from openpyxl import Workbook
from tabulate import tabulate
from categoria_negocio import CategoriaNegocio
from datetime import datetime

listado_categoria=[]
#region categorias
negocio = CategoriaNegocio()

# Funion que imprime en un txt com los datos de los autores
def reporte_categoria():
    print("\n\tGenerando el Reporte de categorias")
    input_excel_file = 'listado_categorias.xlsx'
    # Lee el archivo Excel en un DataFrame
    df = pd.read_excel(input_excel_file)
    # Convierte el DataFrame a una tabla formateada como una cadena de texto
    tabla_formateada = tabulate(df, headers='keys', tablefmt='fancy_outline')
    
    # Sacando fecha para el archivo
    fecha_actual = datetime.now()
    formato = fecha_actual.strftime("%d_%m_%Y")
    print("Fecha actual en formato 'día_mes_año':", formato)
    nom_reporte = 'Reporte_Categorias_'+ formato + '.txt'
    
    # Guarda la tabla formateada en un archivo de texto
    with open(nom_reporte, 'w', encoding='utf-8') as archivo_txt:
        archivo_txt.write(tabla_formateada)


def registrar_categorias():
    codigo_categoria = input('\nIngrese Codigo Categoria: ')
    categoria = input('Ingrese nombre de la Categoria: ')
    negocio.registrar_categorias(codigo_categoria,categoria)
    negocio.guardar_categorias()


def obtener_categorias():
    
    print('\n')
    print('\n \t\tcategorias registradas\n'.center(141,'='))
    # Mostrando todos los categorias registrados
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_categorias.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Codigo de la Categoria', 'Categoria']
    print(tabulate(data,headers=headers,tablefmt='fancy_grid'),'\n')
    # ----------------------------------------------------------------



# Funcion que permite editar categoria
def editar_categoria():
    
    obtener_categorias()
    indice = int(input('\nIngrese el indice del categoria a editar: ')) - 1
    codigo_categoria = int(input('\nIngrese Codigo Categoria: '))
    categoria = input('Ingrese nombre de la Categoria: ')
    
    print(negocio.editar_categorias(indice,codigo_categoria,categoria))


# funcion que permite eliminar categoria
def eliminar_categoria():
    indice_a_eliminar = int(input("\nIngrese el indice del categoria a eliminar: ")) - 1
    print(negocio.eliminar_categoria(indice_a_eliminar))
#endregion



"""
    Categoria(ACT01, 'Acción')
    Categoria(COM02, 'Comedia')
    Categoria(DRM03, 'Drama')
    Categoria(THR04, 'Suspenso')
    Categoria(SCI05, 'Ciencia Ficción')
    Categoria(HOR06, 'Terror')
    Categoria(ROM07, 'Romance')
    Categoria(ANI08, 'Animación')
    Categoria(DOC09, 'Documental')
    Categoria(FAN10, 'Fantasía')
"""


##diccionario
opciones = {
    "1": registrar_categorias,
    "2": obtener_categorias,
    "3": editar_categoria,
    "4": eliminar_categoria,
    "5": reporte_categoria,
    "6": exit
}


while True:
    print("\n\t\tMenú\n".center(80,'='))
    print("1. registrar categorias")
    print("2. obtener categorias")
    print("3. editar categoria")
    print("4. eliminar categoria")
    print("4. reporte categoria (txt)")
    print("5. Salir")
    print("".center(36,'='))
    
    seleccion = input("Seleccione una opción: ")

    if seleccion in opciones:
        opciones[seleccion]()
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")