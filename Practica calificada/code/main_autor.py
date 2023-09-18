# from openpyxl import Workbook, load_workbook
import pandas as pd
import openpyxl
from tabulate import tabulate
from autor_negocio import  AutorNegocio
from datetime import datetime

listado_autores=[]

#region autores
negocio = AutorNegocio()

# Funion que imprime en un txt com los datos de los autores
def reporte_autor():
    print("\n\tGenerando el Reporte de alumno")
    input_excel_file = 'listado_autores.xlsx'
    # Lee el archivo Excel en un DataFrame
    df = pd.read_excel(input_excel_file)
    # Convierte el DataFrame a una tabla formateada como una cadena de texto
    tabla_formateada = tabulate(df, headers='keys', tablefmt='fancy_outline')
    
    # Sacando fecha para el archivo
    fecha_actual = datetime.now()
    formato = fecha_actual.strftime("%d_%m_%Y")
    print("Fecha actual en formato 'día_mes_año':", formato)
    nom_reporte = 'Reporte_Autores_'+ formato + '.txt'
    
    # Guarda la tabla formateada en un archivo de texto
    with open(nom_reporte, 'w', encoding='utf-8') as archivo_txt:
        archivo_txt.write(tabla_formateada)

def registrar_autores():
    nombre = input('Ingrese Nombre: ')
    apellido_paterno = input('Ingrese Apellido Paterno: ')
    apellido_materno = input('Ingrese Apellido Materno: ')
    fecha_nacimiento = input('Ingrese Fecha de Nacimiento: ')
    codigo_autor = input('Ingrese Codigo del Autor: ')
    pais = input('Ingrese Pais: ')
    editorial = input('Ingrese Editorial: ')
    negocio.registrar_autores(nombre, apellido_paterno, apellido_materno, fecha_nacimiento, codigo_autor, pais, editorial)
    negocio.guardar_autores()
    print(f'registro correctamente el autor')

def obtener_autores():
    print('\n')
    print('\n \t\t\t\t\t\t\tAutores registrados\n'.center(313,'='))
    
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_autores.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Nombre', 'Apellido Paterno', 'Apellido Materno', 'Fecha de Nacimiento', 'Codigo del Autor', 'Pais', 'Editorial']
    
    # print(tabulate(data,headers=headers),'\n')
    print(tabulate(data,headers=headers,tablefmt='fancy_grid'),'\n')
    # ----------------------------------------------------------------

def editar_autor():
    
    obtener_autores()
    indice = int(input('Ingrese el indice a editar: '))
    nombre = input('Ingrese Nombre: ')
    apellido_paterno = input('Ingrese Apellido Paterno: ')
    apellido_materno = input('Ingrese Apellido Materno: ')
    fecha_nacimiento = input('Ingrese Fecha de Nacimiento: ')
    codigo_autor = input('Ingrese Codigo del Autor: ')
    pais = input('Ingrese Pais: ')
    editorial = input('Ingrese Editorial: ')
    print(negocio.editar_autor(indice - 1, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, codigo_autor, pais,editorial))

def eliminar_autor():
    #  Muestra los cursos registrados
    obtener_autores()
    
    indice_a_eliminar = int(input("\nIngrese el indice del autor a eliminar: ")) - 1
    print(negocio.eliminar_autor(indice_a_eliminar))
#endregion

"""
    Autor("Jorge", "García", "Pérez", "1980-05-15", "AG1234", "México", "Editorial Alpha"),
    Autor("Ana", "Martínez", "López", "1975-03-20", "BL5678", "España", "Editorial Beta"),
    Autor("David", "Rodríguez", "Sánchez", "1990-08-10", "CA4321", "Argentina", "Editorial Gamma"),
    
    Autor("María", "López", "Fernández", "1988-11-25", "DF9876", "Chile", "Editorial Delta"),
    Autor("Pedro", "Fernández", "Gómez", "1972-02-05", "EG2468", "Perú", "Editorial Epsilon"),
    Autor("Sofía", "González", "Ríos", "1985-06-30", "FR1357", "Colombia", "Editorial Zeta"),
    
    Autor("Carlos", "Hernández", "Molina", "1995-04-15", "GM8642", "Ecuador", "Editorial Eta"),
    Autor("Laura", "Torres", "Paz", "1982-09-12", "HP5555", "Uruguay", "Editorial Theta"),
    Autor("Daniel", "Ríos", "Ramírez", "1979-12-18", "IR1111", "Venezuela", "Editorial Iota"),
    Autor("Isabella", "Silva", "Luna", "1993-07-08", "JL9999", "Costa Rica", "Editorial Kappa")
"""



##diccionario
opciones = {
    "1": registrar_autores,
    "2": obtener_autores,
    "3": editar_autor,
    "4": eliminar_autor,
    "5": reporte_autor,
    "6": exit
}

while True:
    print("\n\t\tMenú\n".center(80,'='))
    print("1. Registrar autores")
    print("2. Listar autores")
    print("3. Editar autor")
    print("4. Eliminar autor")
    print("5. Reporte autor (txt)")
    print("6. Salir")
    print("".center(36,'='))    
    seleccion = input("Seleccione una opción: ")

    if seleccion in opciones:
        opciones[seleccion]()
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")